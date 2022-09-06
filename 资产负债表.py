import requests
import re
from bs4 import BeautifulSoup
from openpyxl import Workbook

china_zfczb=["流动负债","流动资产","非流动资产",'非流动负债','所有者权益(或股东权益)']


# 使用requests库获得文本
stock_code =input("请输入要查询的股票代码:数字即可").strip()
"""选择报告的时间段"""
# 按年度计算
time = 2022
# 按报告期计算
year = 2021     #要查询的年份报告期
month = 6     #要查询最近报告期月份
month_day={"3":"31","6":"30","9":"30","12":"31"}
i = 0
month_list = []
year_list = []
month_list.append(month)
year_list.append(str(year))

while i<4:

    month=month-3

    if month ==0:
        month = 12
        year = year-1

    year_list.append(str(year))
    month_list.append(month)
    i += 1

# print(month_day[str(month_list[0])])

stock_plate = ["sz","SZ","sh","SH"]
"""判断其板块"""

if stock_code.startswith("6"):
    stock_demo = "sh"+stock_code
    stock_demo1 = "SH"+stock_code
elif stock_code.startswith(("0","3")):
    stock_demo = "sz" + stock_code
    stock_demo1 = "SZ" + stock_code
else:
    print("无法判断股票代码:")


url = f"http://emweb.eastmoney.com/PC_HSF10/NewFinanceAnalysis/Index?type=web&code={stock_demo}"
resp = requests.get(url, timeout=40)

"""用beautifulsoup提取信息：股票名称，"""
soup = BeautifulSoup(resp.text,"html.parser")
title_list= soup.find("title").text.split("(")
title=title_list[0]
# print(title_list)













"""用re库正则表达式取得信息"""
# 用re库正则提取资产负债表的脚本代码内容



zcfzb_model = re.compile(r'<script type="text/template" id="zcfzb_qy">.*?</script>'
                         ,re.S)
# 提取tbody标签内容
zcfzb_mode0 = re.compile(r'(<tbody>\s+<tr>\s+<th class="tips-colname-Left" style="width: 366px;">.*?</tbody>)'
                         ,re.S)

#提取name和数据代码
name_model = re.compile(r'\s+([\u4e00-\u9fa5].*?)\s+',re.S)
name1_model = re.compile(r'\s+(其中:\w*?)\s+',re.S)
date_model = re.compile(r"\{\{format\w*?\(value.(?P<date>\w*?)\)}}",re.S)
#使用re提取
zcf =zcfzb_model.findall(resp.text)
zcfzb =''.join(zcf)
zcf1 ="".join(zcfzb_mode0.findall(zcfzb))

print(zcf1)
# print(zcfzb)
#beautifulsoup库提取
soup = BeautifulSoup(zcf1,"html.parser")
trs = soup.find_all("tr")

#创建一个列表，用于储存字典
lis =[]
# print(len(trs))
for tr in trs:
    # 注意字典的添加会有相似代替
    dic = {}
    # print(tr.text)
    names=name_model.findall(tr.text)
    names1=name1_model.findall(tr.text)
    date =date_model.findall(tr.text)
    # print(names,names1)
    if names1:
        # print(names1)
        name_str ="".join(names1)
        dic["name"]=name_str
    else:
        # print(names)
        name_str ="".join(names)
        dic["name"] = name_str
    date_str="".join(date)
    dic["date"]=date_str
    # print(dic)
    # print(date)
    lis.append(dic)
    # print("--------")
lis[0]["date"] = "REPORT_DATE"
# print(lis)
# print(len(lis))



url1 = f"http://emweb.eastmoney.com/PC_HSF10/NewFinanceAnalysis/zcfzbAjaxNew?companyType=4&reportDateType=1&reportType=1&dates={time}-12-31%2C{time-1}-12-31%2C{time-2}-12-31%2C{time-3}-12-31%2C{time-4}-12-31&code={stock_demo1}"
url2 = f"http://emweb.eastmoney.com/PC_HSF10/NewFinanceAnalysis/zcfzbAjaxNew?companyType=4&reportDateType=0&reportType=1&dates={year_list[0]}-{month_list[0]}-{month_day[str(month_list[0])]}%2C{year_list[1]}-{month_list[1]}-{month_day[str(month_list[1])]}%2C{year_list[2]}-{month_list[2]}-{month_day[str(month_list[2])]}%2C{year_list[3]}-{month_list[3]}-{month_day[str(month_list[3])]}%2C{year_list[4]}-{month_list[4]}-{month_day[str(month_list[4])]}&code={stock_demo1}"

resp1 = requests.get(url2, timeout=40)

# print(resp1.text)
resp_json=resp1.json()
# print(resp_json)
date_list =resp_json["data"]
count = resp_json["count"]
# print(count)
for i in date_list:
    # 获得遍历列表的元素序号,col是字符串
    col = str(date_list.index(i)+1)

    for n in lis:
        # 获得当前得到的是哪个字典的数据
        # print(n)
        index =lis.index(n)
        date = n["date"]
        try:
            info = i[date]
            # print(info)
            # print(type(info))
            lis[index][col] = info
            # print("一个项目的数据完成了\n")


        except:
            if n["name"]=="吸收存款及同业存放":
              info = ""
              lis[index][col] = info
            else:
                info = "--"
                # print(info)
                # print(type(info))
                lis[index][col] = info
                # print("一个项目的数据完成了---\n")
        # print("一年的数据成功的完成了\n")
# print(lis)






"""创建一个Excel工作簿，为xlsx格式"""
wb = Workbook()
# 新建一个工作表并命名，
ws = wb.active
ws.title = "资产负债表1"

# 计算表的行数和列数
rows = len(lis)
cols = len(lis[0])



for a in lis:
    row_n = lis.index(a)
    # print(row_n)
    col= 0
    for v in a.values():
        if col<cols:
            ws.cell(row_n+1,col+1,v)
            col += 1
        else:
            print("出错")
# 删除不要的数据，以便美化图表
ws.delete_cols(2)   #删除数据data列
max_row=ws.max_row
max_col = ws.max_column
col_list = [j for j in range(max_col)]
# print(max_row,man_col)
shanchu =[ ]
for r in range(max_row):
    # 数量限制判断
    x = 0
    for c in range(max_col):
        if ws.cell(row=r+1, column=c+1).value in [None,"",0,"0"]:
            x += 1
            # print("x",x)
            if x >4:
                #记录要删除的行
                shanchu.append(r+1)
            else:
                continue
        else:
            continue
# print(shanchu)
# print(len(shanchu))
ws1 = wb.create_sheet("资产负债表")
dat = [i for i in range(1,len(lis)+1)]

for c in shanchu:
    dat.remove(c)
# print(dat)
for z in dat:
    q=dat.index(z)+1
    for c in col_list:
        r = col_list.index(c) + 1
        ws1.cell(q,r,ws.cell(z,r).value)

wb.remove(ws)







# row_number = 0
# for row in ws.iter_rows():#迭代遍历每行
#     if cols[2,6] is None:   #给定的条件，读者可根据自身需求自定义
#         row_number=col[2:6].row #关键步骤！获得当前行的行号！
#         print(row_number)
#         ws.delete_rows(row_number)






# lis1=[v for a in lis for v in a.values()]
# print(lis1)
wb.save("东方财富网/"+f"{title}.xlsx")
print('保存成功')





















