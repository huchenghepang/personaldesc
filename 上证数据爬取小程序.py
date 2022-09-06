from selenium import webdriver
from openpyxl import Workbook
from selenium.webdriver.chrome.options import Options
from time import sleep,time


import time


def merge_url(stock_id,report_period):
    """处理拼接url"""
    url = f"http://listxbrl.sse.com.cn/companyInfo/toCompanyInfo.do?stock_id=" \
          f"{stock_id}&report_period_id={report_period}"
    return url



def open_web(url):
    """使用selenium打开网页,并等待一会儿,如需要请调整等待时间"""
    options = Options()
    options.add_argument('-headless')
    options.add_argument('--disable-gpu')
    wd = webdriver.Chrome(options=options)
    wd.get(url)
    sleep(1)
    return wd

def acquire_sheet_element(wd):
    """获取表头元素"""
    lis_head = []
    head_tr = wd.find_element_by_xpath(r'//*[@id="index_center"]/div[1]/div/div/div[2]/div[1]/div/table/tbody/tr')
    head_tds = head_tr.find_elements_by_xpath('.//td')
    for td in head_tds:
        head_text = td.find_element_by_xpath('.//div/span').text
        lis_head.append(head_text)
    return lis_head


def create_sheet():
    """创建工作表"""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "a"
    sheet_name = ["基本信息", "股本结构", "前十大股东", "资产负债表", "利润表", "现金流量表"]
    return sheet_name,wb


def workbook_name(wd):
    """文件名"""
    book_name = wd.find_element_by_xpath('/html/body/div[4]/div/div[2]/div[4]/div/ul/li[7]')
    book_name = book_name.text[5:-7]
    return book_name

def final_inf(wd,lis_head):
    # 建立表名
    sheet_name,wb =create_sheet()
    # 依次点击要查找的表,注意elements
    ul_element = wd.find_element_by_xpath(r'//*[@id="tabs"]')
    lis = ul_element.find_elements_by_xpath('.//li/a')
    """循环遍历相关标签和路径"""
    for li in lis:
        li_num = lis.index(li)
        sheet = wb.create_sheet(f'{sheet_name[li_num]}')
        sheet.append(lis_head)
        li.click()
        sleep(1)
        tbody = wd.find_element_by_xpath('//*[@id="index_center"]/div[1]/div/div/div[2]/div[2]/table/tbody')
        # sleep(2)
        """获取表的其他部分"""
        trs = tbody.find_elements_by_xpath('.//tr')
        for tr in trs:
            row = trs.index(tr)
            # print(row)
            tds = tr.find_elements_by_xpath('.//td')
            lis1 = []
            for t in tds:
                col = tds.index(t)
                data_text = t.find_element_by_xpath('.//div').text
                # print(data_text)
                sheet.cell(row + 2, col + 1, data_text)
        """得到一个表"""
        print("--------------------")
    del wb["a"]
    return wb


def main():
    """输入要查询的股票代码和公司简称，以及报告期{"年报"：5000，"半年报"：1000，"一季度"：4000，"三季度"：4400}"""
    stock_id = 600361
    # 输入对应值
    report_period = 5000
    url = merge_url(stock_id,report_period)
    wd = open_web(url)
    lis_head = acquire_sheet_element(wd)
    # sheet_name,wb = create_sheet()
    book_name= workbook_name(wd)
    wb = final_inf(wd,lis_head)
    filename = book_name + str(stock_id)+".xlsx"
    wb.save(filename)
    print("保存成功")


# start_time =time.time()
main()
# end_time = time.time()
# time = end_time-start_time
# print("花费的时间:",time,"s")







